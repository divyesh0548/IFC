#!/usr/bin/env python3
"""
Merge all Excel files in a folder into one workbook.

Columns are the union of headers across files (first-seen order).
If a file does not have a column that exists in another file, that cell
is left blank for that file's rows.

Run:
    pip install openpyxl
    python merge_files.py

Or:
    python merge_files.py "C:\\path\\to\\folder"
    python merge_files.py "C:\\path\\to\\folder" -o merged.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION — used when no CLI folder is provided
# ============================================================

INPUT_FOLDER = r"C:\Divyesh\IFC\IFC_Prisma\Controls Library\CAPEX"
OUTPUT_FILE_NAME = "merged_controls.xlsx"
# First worksheet only. Set True to append every sheet from each file.
INCLUDE_ALL_SHEETS = False

# ============================================================

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
HEADER_SCAN_ROWS = 30


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def merged_top_left(worksheet, row_number: int, column_number: int) -> tuple[int, int]:
    cell = worksheet.cell(row=row_number, column=column_number)
    if not isinstance(cell, MergedCell):
        return row_number, column_number

    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, _max_col, _max_row = merged_range.bounds
            return min_row, min_col
    return row_number, column_number


def cell_value(worksheet, row_number: int, column_number: int) -> Any:
    source_row, source_col = merged_top_left(worksheet, row_number, column_number)
    return worksheet.cell(row=source_row, column=source_col).value


def row_width(worksheet, row_number: int) -> int:
    last_column = 0
    max_column = worksheet.max_column or 1
    for column_number in range(1, max_column + 1):
        value = cell_value(worksheet, row_number, column_number)
        if value is not None and str(value).strip():
            last_column = column_number
    return last_column


def detect_header_row(worksheet) -> int:
    """Prefer the densest non-empty early row as the header."""
    best_row = 1
    best_count = -1
    max_scan = min(worksheet.max_row or 1, HEADER_SCAN_ROWS)

    for row_number in range(1, max_scan + 1):
        width = row_width(worksheet, row_number)
        count = 0
        for column_number in range(1, width + 1):
            value = cell_value(worksheet, row_number, column_number)
            if value is not None and str(value).strip():
                count += 1
        if count > best_count:
            best_count = count
            best_row = row_number

    return best_row


def read_headers(worksheet, header_row: int) -> list[str]:
    width = max(row_width(worksheet, header_row), 1)
    headers: list[str] = []
    seen: dict[str, int] = {}

    for column_number in range(1, width + 1):
        raw = cell_value(worksheet, header_row, column_number)
        name = normalize_header(raw)
        if not name:
            name = f"Column_{get_column_letter(column_number)}"

        # Keep duplicate headers unique within a single sheet.
        key = name.casefold()
        if key in seen:
            seen[key] += 1
            name = f"{name}_{seen[key]}"
        else:
            seen[key] = 1
        headers.append(name)

    return headers


def read_data_rows(worksheet, header_row: int, headers: list[str]) -> list[dict[str, Any]]:
    width = len(headers)
    last_row = header_row
    max_row = worksheet.max_row or header_row

    for row_number in range(header_row + 1, max_row + 1):
        for column_number in range(1, width + 1):
            value = cell_value(worksheet, row_number, column_number)
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            last_row = row_number
            break

    rows: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, last_row + 1):
        record: dict[str, Any] = {}
        empty = True
        for column_number, header in enumerate(headers, start=1):
            value = cell_value(worksheet, row_number, column_number)
            if value is not None and not (isinstance(value, str) and not value.strip()):
                empty = False
            record[header] = value
        if not empty:
            rows.append(record)
    return rows


def list_excel_files(folder: Path) -> list[Path]:
    files = [
        path
        for path in sorted(folder.iterdir(), key=lambda item: item.name.casefold())
        if path.is_file()
        and path.suffix.lower() in EXCEL_SUFFIXES
        and not path.name.startswith("~$")
    ]
    return files


def choose_folder_dialog() -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askdirectory(title="Select folder containing Excel files")
    root.destroy()
    if not selected:
        return None
    return Path(selected)


def merge_excel_folder(
    folder: Path,
    output_path: Path,
    include_all_sheets: bool = False,
) -> dict[str, Any]:
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {folder}")

    excel_files = list_excel_files(folder)
    if not excel_files:
        raise ValueError(f"No Excel files (.xlsx / .xlsm) found in: {folder}")

    # Skip overwriting the merge target if it already sits in the same folder.
    excel_files = [path for path in excel_files if path.resolve() != output_path.resolve()]
    if not excel_files:
        raise ValueError("No Excel files left to merge after excluding the output file.")

    all_columns: list[str] = []
    column_keys: set[str] = set()
    merged_rows: list[dict[str, Any]] = []
    file_summaries: list[str] = []

    for file_path in excel_files:
        workbook = load_workbook(filename=file_path, data_only=False, read_only=False)
        sheet_names = list(workbook.sheetnames) if include_all_sheets else workbook.sheetnames[:1]
        file_row_count = 0

        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            if (worksheet.max_row or 0) < 1:
                continue

            header_row = detect_header_row(worksheet)
            headers = read_headers(worksheet, header_row)
            if not headers:
                continue

            for header in headers:
                key = header.casefold()
                if key not in column_keys:
                    column_keys.add(key)
                    all_columns.append(header)

            # Map this sheet's headers onto the global names (case-insensitive).
            local_to_global = {}
            global_by_key = {name.casefold(): name for name in all_columns}
            for header in headers:
                local_to_global[header] = global_by_key[header.casefold()]

            for record in read_data_rows(worksheet, header_row, headers):
                merged: dict[str, Any] = {column: None for column in all_columns}
                for local_header, value in record.items():
                    global_header = local_to_global[local_header]
                    merged[global_header] = value
                # Ensure later-discovered columns stay present for earlier rows.
                for column in all_columns:
                    merged.setdefault(column, None)
                merged_rows.append(merged)
                file_row_count += 1

        workbook.close()
        file_summaries.append(f"{file_path.name}: {file_row_count} row(s)")

    # Backfill newly discovered columns into earlier rows.
    for row in merged_rows:
        for column in all_columns:
            row.setdefault(column, None)

    if not all_columns:
        raise ValueError("No header columns were found in the selected Excel files.")

    output = Workbook()
    sheet = output.active
    sheet.title = "Merged"

    for index, column in enumerate(all_columns, start=1):
        sheet.cell(row=1, column=index, value=column)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(column) + 4, 14), 48)

    for row_index, record in enumerate(merged_rows, start=2):
        for column_index, column in enumerate(all_columns, start=1):
            sheet.cell(row=row_index, column=column_index, value=record.get(column))

    sheet.freeze_panes = "A2"
    if merged_rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}{len(merged_rows) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)

    return {
        "files": len(excel_files),
        "columns": len(all_columns),
        "rows": len(merged_rows),
        "output": output_path,
        "summaries": file_summaries,
        "column_names": all_columns,
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge all Excel files in a folder into one file with a union of columns.",
    )
    parser.add_argument(
        "folder",
        nargs="?",
        default="",
        help="Folder containing Excel files. If omitted, a folder picker opens.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="",
        help="Output Excel path or file name. Default: merged_controls.xlsx in the folder.",
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Include every worksheet from each Excel file (default: first sheet only).",
    )
    return parser.parse_args(argv)


def resolve_paths(args: argparse.Namespace) -> tuple[Path, Path, bool]:
    folder_text = str(args.folder or INPUT_FOLDER or "").strip()
    if folder_text:
        folder = Path(folder_text).expanduser().resolve()
    else:
        picked = choose_folder_dialog()
        if picked is None:
            raise SystemExit("No folder selected.")
        folder = picked.resolve()

    output_text = str(args.output or OUTPUT_FILE_NAME or "merged_controls.xlsx").strip()
    output_path = Path(output_text).expanduser()
    if not output_path.is_absolute():
        output_path = folder / output_path
    if output_path.suffix.lower() not in EXCEL_SUFFIXES:
        output_path = output_path.with_suffix(".xlsx")

    include_all_sheets = bool(args.all_sheets or INCLUDE_ALL_SHEETS)
    return folder, output_path.resolve(), include_all_sheets


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        folder, output_path, include_all_sheets = resolve_paths(args)
        result = merge_excel_folder(
            folder=folder,
            output_path=output_path,
            include_all_sheets=include_all_sheets,
        )
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    print(f"Folder: {folder}")
    print(f"Files merged: {result['files']}")
    print(f"Columns: {result['columns']}")
    print(f"Rows: {result['rows']}")
    print("Per-file counts:")
    for line in result["summaries"]:
        print(f"  - {line}")
    print(f"\nSaved: {result['output']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
