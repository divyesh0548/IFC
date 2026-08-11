#!/usr/bin/env python3
"""
Remap Excel control-library columns to the required set and drop unused columns.

Run:
    pip install -r requirements.txt
    python main.py
"""

from __future__ import annotations

import re
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = [
    "Sub-Process",
    "Risk Description",
    "Risk Heat",
    "Control Objective",
    "Standard Control Description",
    "Control type (Manual/Automated)",
    "Control type (Financial/Operational)",
    "Nature of Control (Preventive/Detective)",
    "Process Activity and Walkthrough details",
    "Key Control",
    "Application name",
    "Control Evidence to be obtained",
    "Whether fraud risk exists? (Yes/No)",
    "Control Frequency",
]

UNMAPPED_LABEL = "— Not mapped —"

# Phrase / keyword groups used for header detection and auto-mapping.
COLUMN_KEYWORDS: dict[str, list[list[str]]] = {
    "Sub-Process": [
        ["sub", "process"],
        ["subprocess"],
        ["other", "affected"],
    ],
    "Risk Description": [
        ["risk", "description"],
        ["risk", "desc"],
        ["risk"],
    ],
    "Risk Heat": [
        ["risk", "heat"],
        ["heat"],
    ],
    "Control Objective": [
        ["control", "objective"],
        ["objective"],
    ],
    "Standard Control Description": [
        ["standard", "control", "description"],
        ["control", "description"],
        ["control", "name"],
    ],
    "Control type (Manual/Automated)": [
        ["manual", "automated"],
        ["type", "manual"],
        ["type", "automated"],
    ],
    "Control type (Financial/Operational)": [
        ["financial", "operational"],
        ["type", "financial"],
        ["type", "operational"],
    ],
    "Nature of Control (Preventive/Detective)": [
        ["nature", "control"],
        ["preventive", "detective"],
        ["preventive"],
        ["detective"],
    ],
    "Process Activity and Walkthrough details": [
        ["process", "activity"],
        ["walkthrough"],
        ["process", "walkthrough"],
    ],
    "Key Control": [
        ["key", "control"],
    ],
    "Application name": [
        ["application", "name"],
        ["application"],
    ],
    "Control Evidence to be obtained": [
        ["control", "evidence"],
        ["audit", "evidence"],
        ["evidence", "obtained"],
        ["evidence"],
    ],
    "Whether fraud risk exists? (Yes/No)": [
        ["whether", "fraud"],
        ["fraud", "risk"],
        ["fraud"],
    ],
    "Control Frequency": [
        ["control", "frequency"],
        ["frequency"],
    ],
}

HEADER_SCAN_ROWS = 40
MIN_HEADER_HITS = 3
AUTO_MAP_MIN_SCORE = 8


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def display_header(value: Any, column_number: int) -> str:
    text = "" if value is None else str(value).strip()
    if text:
        return text
    return f"(Blank column {get_column_letter(column_number)})"


def keyword_match_score(normalized_header: str, tokens: set[str], keyword_group: list[str]) -> int:
    if not keyword_group:
        return 0

    phrase = " ".join(keyword_group)
    if phrase == normalized_header:
        return 100 + (10 * len(keyword_group))
    if phrase and phrase in f" {normalized_header} ":
        return 70 + (8 * len(keyword_group))

    if all(word in tokens for word in keyword_group):
        return 40 + (12 * len(keyword_group))

    hits = sum(1 for word in keyword_group if word in tokens)
    if hits == 0:
        return 0
    if len(keyword_group) == 1:
        return 6 if hits else 0
    return hits * 4


def score_header_against_target(header_value: Any, target: str) -> int:
    normalized = normalize_text(header_value)
    if not normalized:
        return 0

    tokens = set(normalized.split())
    best = 0
    for group in COLUMN_KEYWORDS.get(target, []):
        best = max(best, keyword_match_score(normalized, tokens, group))

    # Avoid mapping generic "risk" onto heat / fraud columns, and vice versa.
    if target == "Risk Description":
        if "heat" in tokens or "fraud" in tokens:
            best = min(best, 5)
    if target == "Risk Heat" and "heat" not in tokens:
        best = 0
    if target == "Sub-Process" and "sub" not in tokens and "subprocess" not in tokens:
        if "other" not in tokens:
            best = min(best, 4)
    if target == "Control Frequency" and "frequency" not in tokens:
        best = 0
    if target == "Key Control" and "key" not in tokens:
        best = 0

    return best


def detect_header_row(worksheet) -> tuple[int, int]:
    """Return (header_row, used_column_count)."""
    best_row = 1
    best_score = -1
    best_width = 1
    max_scan_row = min(worksheet.max_row or 1, HEADER_SCAN_ROWS)

    for row_number in range(1, max_scan_row + 1):
        values = read_row_values(worksheet, row_number)
        non_empty = [value for value in values if str(value or "").strip()]
        if len(non_empty) < 2:
            continue

        hits = 0
        for value in non_empty:
            if any(score_header_against_target(value, target) >= AUTO_MAP_MIN_SCORE for target in REQUIRED_COLUMNS):
                hits += 1

        if hits > best_score or (hits == best_score and len(non_empty) > best_width):
            best_score = hits
            best_row = row_number
            best_width = len(values)

    if best_score < MIN_HEADER_HITS:
        # Fall back to the densest early row so the user can still map manually.
        densest_row = 1
        densest_count = 0
        for row_number in range(1, max_scan_row + 1):
            values = read_row_values(worksheet, row_number)
            count = sum(1 for value in values if str(value or "").strip())
            if count > densest_count:
                densest_count = count
                densest_row = row_number
        return densest_row, max(read_row_width(worksheet, densest_row), 1)

    return best_row, max(best_width, 1)


def merged_top_left(worksheet, row_number: int, column_number: int) -> tuple[int, int]:
    cell = worksheet.cell(row=row_number, column=column_number)
    if not isinstance(cell, MergedCell):
        return row_number, column_number

    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, _max_col, _max_row = merged_range.bounds
            return min_row, min_col
    return row_number, column_number


def cell_value(worksheet, row_number: int, column_number: int) -> Any:
    source_row, source_col = merged_top_left(worksheet, row_number, column_number)
    return worksheet.cell(row=source_row, column=source_col).value


def read_row_width(worksheet, row_number: int) -> int:
    last_column = 0
    max_column = worksheet.max_column or 1
    for column_number in range(1, max_column + 1):
        value = cell_value(worksheet, row_number, column_number)
        if value is not None and str(value).strip():
            last_column = column_number
    return last_column


def read_row_values(worksheet, row_number: int) -> list[Any]:
    width = max(read_row_width(worksheet, row_number), worksheet.max_column or 1)
    return [cell_value(worksheet, row_number, column_number) for column_number in range(1, width + 1)]


def auto_map_columns(headers: list[tuple[int, Any]]) -> dict[int, str]:
    """Return {source_column_number: required_column_name}."""
    candidates: list[tuple[int, int, str]] = []
    for column_number, header_value in headers:
        for target in REQUIRED_COLUMNS:
            score = score_header_against_target(header_value, target)
            if score >= AUTO_MAP_MIN_SCORE:
                candidates.append((score, column_number, target))

    candidates.sort(key=lambda item: (-item[0], item[1], REQUIRED_COLUMNS.index(item[2])))
    mapping: dict[int, str] = {}
    used_targets: set[str] = set()
    used_sources: set[int] = set()

    for _score, column_number, target in candidates:
        if column_number in used_sources or target in used_targets:
            continue
        mapping[column_number] = target
        used_sources.add(column_number)
        used_targets.add(target)

    return mapping


def last_used_row(worksheet, header_row: int, column_count: int) -> int:
    last_row = header_row
    for row_number in range(header_row + 1, (worksheet.max_row or header_row) + 1):
        if any(
            cell_value(worksheet, row_number, column_number) not in (None, "")
            for column_number in range(1, column_count + 1)
        ):
            last_row = row_number
    return last_row


def default_output_path(source_path: Path, custom_name: str) -> Path:
    name = custom_name.strip()
    if not name:
        name = source_path.stem
    if not name.lower().endswith((".xlsx", ".xlsm")):
        name = f"{name}{source_path.suffix.lower() or '.xlsx'}"
    return source_path.with_name(name)


def write_normalized_workbook(
    source_worksheet,
    header_row: int,
    mapping: dict[int, str],
    destination: Path,
) -> int:
    ordered_targets = [target for target in REQUIRED_COLUMNS if target in mapping.values()]
    source_by_target = {target: column for column, target in mapping.items()}

    output = Workbook()
    sheet = output.active
    sheet.title = (source_worksheet.title or "Controls")[:31]

    for index, target in enumerate(ordered_targets, start=1):
        sheet.cell(row=1, column=index, value=target)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(target) + 4, 18), 48)

    column_count = max(mapping.keys(), default=1)
    data_last_row = last_used_row(source_worksheet, header_row, column_count)
    output_row = 1
    copied_rows = 0

    for source_row in range(header_row + 1, data_last_row + 1):
        values = [
            cell_value(source_worksheet, source_row, source_by_target[target])
            for target in ordered_targets
        ]
        if all(value in (None, "") or (isinstance(value, str) and not value.strip()) for value in values):
            continue
        output_row += 1
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=output_row, column=column_index, value=value)
        copied_rows += 1

    sheet.freeze_panes = "A2"
    if ordered_targets:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(ordered_targets))}{max(output_row, 1)}"

    output.save(destination)
    return copied_rows


class ColumnNormalizerApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Control Column Normalizer")
        self.root.geometry("980x720")
        self.root.minsize(820, 560)

        self.source_path: Path | None = None
        self.workbook = None
        self.worksheet = None
        self.header_row = 1
        self.headers: list[tuple[int, Any]] = []
        self.mapping_vars: dict[int, tk.StringVar] = {}
        self.mapping_combos: dict[int, ttk.Combobox] = {}

        self._build_ui()

    def _build_ui(self) -> None:
        padding = {"padx": 12, "pady": 8}
        outer = ttk.Frame(self.root, padding=16)
        outer.pack(fill=tk.BOTH, expand=True)

        ttk.Label(outer, text="Control Column Normalizer", font=("Segoe UI", 16, "bold")).pack(anchor="w")
        ttk.Label(
            outer,
            text="Upload an Excel file, confirm header mapping, and export only the required control columns.",
        ).pack(anchor="w", pady=(0, 10))

        file_row = ttk.Frame(outer)
        file_row.pack(fill=tk.X, **padding)
        ttk.Button(file_row, text="Upload Excel file", command=self.choose_file).pack(side=tk.LEFT)
        self.file_label = ttk.Label(file_row, text="No file selected")
        self.file_label.pack(side=tk.LEFT, padx=10)

        sheet_row = ttk.Frame(outer)
        sheet_row.pack(fill=tk.X, padx=12)
        ttk.Label(sheet_row, text="Sheet:").pack(side=tk.LEFT)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(sheet_row, textvariable=self.sheet_var, state="disabled", width=40)
        self.sheet_combo.pack(side=tk.LEFT, padx=8)
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda _event: self.load_selected_sheet())

        self.header_info = ttk.Label(outer, text="")
        self.header_info.pack(anchor="w", padx=12, pady=(8, 4))

        table_frame = ttk.LabelFrame(outer, text="Column mapping", padding=8)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        header_bar = ttk.Frame(table_frame)
        header_bar.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(header_bar, text="Source column", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(4, 0))
        ttk.Label(header_bar, text="Required column", font=("Segoe UI", 9, "bold")).pack(side=tk.RIGHT, padx=(0, 180))

        canvas_holder = ttk.Frame(table_frame)
        canvas_holder.pack(fill=tk.BOTH, expand=True)
        self.canvas = tk.Canvas(canvas_holder, highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_holder, orient=tk.VERTICAL, command=self.canvas.yview)
        self.mapping_frame = ttk.Frame(self.canvas)
        self.mapping_frame.bind(
            "<Configure>",
            lambda event: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        self.canvas.create_window((0, 0), window=self.mapping_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        output_row = ttk.Frame(outer)
        output_row.pack(fill=tk.X, padx=12, pady=(4, 0))
        ttk.Label(output_row, text="Output file name (optional):").pack(side=tk.LEFT)
        self.output_name_var = tk.StringVar()
        ttk.Entry(output_row, textvariable=self.output_name_var, width=42).pack(side=tk.LEFT, padx=8)
        ttk.Label(output_row, text="Leave blank to use the original file name.").pack(side=tk.LEFT)

        action_row = ttk.Frame(outer)
        action_row.pack(fill=tk.X, padx=12, pady=12)
        ttk.Button(action_row, text="Generate Excel", command=self.generate_file).pack(side=tk.LEFT)
        self.status_label = ttk.Label(action_row, text="")
        self.status_label.pack(side=tk.LEFT, padx=12)

    def _on_mousewheel(self, event: tk.Event) -> None:
        self.canvas.yview_scroll(int(-event.delta / 120), "units")

    def choose_file(self) -> None:
        selected = filedialog.askopenfilename(
            title="Select control Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not selected:
            return

        path = Path(selected)
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            messagebox.showerror("Unsupported file", "Please select an .xlsx or .xlsm file.")
            return

        try:
            workbook = load_workbook(filename=path, data_only=False, read_only=False)
        except Exception as error:
            messagebox.showerror("Could not open file", str(error))
            return

        self.source_path = path
        self.workbook = workbook
        self.file_label.config(text=str(path))
        self.output_name_var.set(path.name)
        self.sheet_combo.configure(state="readonly", values=workbook.sheetnames)
        self.sheet_var.set(workbook.sheetnames[0] if workbook.sheetnames else "")
        self.load_selected_sheet()

    def load_selected_sheet(self) -> None:
        if self.workbook is None or not self.sheet_var.get():
            return

        self.worksheet = self.workbook[self.sheet_var.get()]
        self.header_row, _width = detect_header_row(self.worksheet)
        width = max(read_row_width(self.worksheet, self.header_row), 1)
        self.headers = [
            (column_number, cell_value(self.worksheet, self.header_row, column_number))
            for column_number in range(1, width + 1)
            if str(cell_value(self.worksheet, self.header_row, column_number) or "").strip()
        ]

        auto_mapping = auto_map_columns(self.headers)
        mapped_count = len(auto_mapping)
        self.header_info.config(
            text=(
                f'Detected header row {self.header_row} on sheet "{self.worksheet.title}". '
                f"{len(self.headers)} source column(s), {mapped_count} auto-mapped."
            )
        )
        self._render_mapping_rows(auto_mapping)
        self.status_label.config(text="Review the auto-detected mapping, then generate.")

    def _render_mapping_rows(self, auto_mapping: dict[int, str]) -> None:
        for child in self.mapping_frame.winfo_children():
            child.destroy()
        self.mapping_vars = {}
        self.mapping_combos = {}

        for column_number, header_value in self.headers:
            row = ttk.Frame(self.mapping_frame)
            row.pack(fill=tk.X, pady=3)

            source_text = display_header(header_value, column_number)
            ttk.Label(row, text=f"{get_column_letter(column_number)}.  {source_text}").pack(
                side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 12)
            )

            current = tk.StringVar(value=auto_mapping.get(column_number, UNMAPPED_LABEL))
            combo = ttk.Combobox(row, textvariable=current, state="readonly", width=42)
            combo.pack(side=tk.RIGHT)
            combo.bind(
                "<<ComboboxSelected>>",
                lambda _event, source_column=column_number: self._on_mapping_changed(source_column),
            )
            self.mapping_vars[column_number] = current
            self.mapping_combos[column_number] = combo

        self._refresh_mapping_choices()
        self.mapping_frame.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_mapping_changed(self, source_column: int) -> None:
        selected = self.mapping_vars[source_column].get().strip()
        if selected and selected != UNMAPPED_LABEL:
            for other_column, variable in self.mapping_vars.items():
                if other_column != source_column and variable.get().strip() == selected:
                    variable.set(UNMAPPED_LABEL)
        self._refresh_mapping_choices()

    def _refresh_mapping_choices(self) -> None:
        used_targets = {
            variable.get().strip()
            for variable in self.mapping_vars.values()
            if variable.get().strip() and variable.get().strip() != UNMAPPED_LABEL
        }

        for column_number, combo in self.mapping_combos.items():
            current = self.mapping_vars[column_number].get().strip() or UNMAPPED_LABEL
            available = [UNMAPPED_LABEL] + [
                target for target in REQUIRED_COLUMNS
                if target == current or target not in used_targets
            ]
            combo.configure(values=available)

    def current_mapping(self) -> dict[int, str]:
        mapping: dict[int, str] = {}
        for column_number, variable in self.mapping_vars.items():
            target = variable.get().strip()
            if target and target != UNMAPPED_LABEL:
                mapping[column_number] = target
        return mapping

    def generate_file(self) -> None:
        if self.source_path is None or self.worksheet is None:
            messagebox.showwarning("No file", "Upload an Excel file first.")
            return

        mapping = self.current_mapping()
        if not mapping:
            messagebox.showwarning("No mapping", "Map at least one source column to a required column.")
            return

        targets = list(mapping.values())
        duplicates = sorted({name for name in targets if targets.count(name) > 1})
        if duplicates:
            messagebox.showerror(
                "Duplicate mapping",
                "Each required column can be mapped only once:\n- " + "\n- ".join(duplicates),
            )
            return

        unmapped_required = [name for name in REQUIRED_COLUMNS if name not in mapping.values()]
        if unmapped_required:
            proceed = messagebox.askyesno(
                "Some required columns are not mapped",
                "These required columns will be omitted from the output:\n- "
                + "\n- ".join(unmapped_required)
                + "\n\nContinue anyway?",
            )
            if not proceed:
                return

        suggested = default_output_path(self.source_path, self.output_name_var.get())
        destination = filedialog.asksaveasfilename(
            title="Save normalized Excel file",
            defaultextension=".xlsx",
            initialdir=str(suggested.parent),
            initialfile=suggested.name,
            filetypes=[("Excel files", "*.xlsx *.xlsm")],
        )
        if not destination:
            return

        try:
            copied_rows = write_normalized_workbook(
                self.worksheet,
                self.header_row,
                mapping,
                Path(destination),
            )
        except Exception as error:
            messagebox.showerror("Could not generate file", str(error))
            return

        self.status_label.config(text=f"Created {Path(destination).name} ({copied_rows} data row(s)).")
        messagebox.showinfo(
            "File generated",
            f"Saved {copied_rows} row(s) to:\n{destination}",
        )


def main() -> None:
    root = tk.Tk()
    try:
        style = ttk.Style(root)
        if "vista" in style.theme_names():
            style.theme_use("vista")
        elif "clam" in style.theme_names():
            style.theme_use("clam")
    except tk.TclError:
        pass

    ColumnNormalizerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
