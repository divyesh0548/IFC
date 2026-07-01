#!/usr/bin/env python3
"""
Split an Excel worksheet into separate Excel files based on distinct values
in a selected column.

Before running, update the variables under CONFIGURATION below.

Dependency:
    pip install openpyxl
"""

from __future__ import annotations

import re
from collections import OrderedDict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


# ============================================================
# CONFIGURATION — CHANGE ONLY THESE VALUES
# ============================================================

INPUT_FILE = r"C:\Divyesh\IFC\IFC_Prisma\Risk Categorization\Base IFCs\TIL\TIL_FY-2025-26_IFC Review.xlsx"
SHEET_NAME = "IFC"
COLUMN_NAME = "Process"
OUTPUT_FOLDER = r"C:\Divyesh\IFC\IFC_Prisma\Risk Categorization\TIL"

# Set True to create Blank.xlsx for rows where the selected column is blank.
INCLUDE_BLANK_VALUES = False

# ============================================================


INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1F]')

WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def normalize_text(value: Any) -> str:
    """Normalize text for case-insensitive and space-insensitive comparison."""
    if value is None:
        return ""

    return " ".join(str(value).strip().casefold().split())


def value_key(value: Any) -> tuple[str, Any]:
    """
    Create a type-aware dictionary key.

    This keeps numeric 1 and text "1" as separate distinct values.
    """
    if isinstance(value, datetime):
        return ("datetime", value.isoformat())

    if isinstance(value, date):
        return ("date", value.isoformat())

    return (type(value).__name__, value)


def value_for_filename(value: Any) -> str:
    """Convert a distinct Excel value into readable filename text."""
    if value is None:
        return "Blank"

    if isinstance(value, str) and not value.strip():
        return "Blank"

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d_%H-%M-%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def create_safe_filename(value: Any, maximum_length: int = 150) -> str:
    """Remove characters that are not permitted in Windows filenames."""
    filename = value_for_filename(value)

    filename = INVALID_FILENAME_CHARS.sub("_", filename)
    filename = re.sub(r"\s+", " ", filename).strip(" .")

    if not filename:
        filename = "Blank"

    if filename.upper() in WINDOWS_RESERVED_NAMES:
        filename = f"_{filename}"

    filename = filename[:maximum_length].rstrip(" .")

    return filename or "Blank"


def create_unique_filename(base_name: str, used_names: set[str]) -> str:
    """
    Prevent files from being overwritten where two values result
    in the same sanitized filename.
    """
    filename = base_name
    number = 2

    while filename.casefold() in used_names:
        filename = f"{base_name}_{number}"
        number += 1

    used_names.add(filename.casefold())
    return filename


def find_header(worksheet, required_column_name: str) -> tuple[int, int]:
    """
    Search the entire selected worksheet for the specified column heading.

    Returns:
        Header row number and selected column number.
    """
    required_header = normalize_text(required_column_name)
    matches: list[tuple[int, int]] = []

    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            if normalize_text(cell.value) == required_header:
                matches.append((cell.row, cell.column))

    if not matches:
        raise ValueError(
            f'Column heading "{required_column_name}" was not found '
            f'in worksheet "{worksheet.title}".'
        )

    if len(matches) > 1:
        cell_addresses = ", ".join(
            f"{get_column_letter(column_number)}{row_number}"
            for row_number, column_number in matches
        )

        raise ValueError(
            f'Column heading "{required_column_name}" was found more than once '
            f'at: {cell_addresses}. Make the heading unique.'
        )

    return matches[0]


def find_last_used_column(worksheet, header_row: int) -> int:
    """Find the final used column from the header row onward."""
    last_used_column = 0

    for row in worksheet.iter_rows(
        min_row=header_row,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if not isinstance(cell, MergedCell) and cell.value is not None:
                last_used_column = max(last_used_column, cell.column)

    return max(last_used_column, 1)


def copy_cell(source_cell, destination_cell) -> None:
    """Copy a cell's value and formatting."""
    destination_cell.value = source_cell.value

    if source_cell.has_style:
        destination_cell._style = copy(source_cell._style)

    if source_cell.number_format:
        destination_cell.number_format = source_cell.number_format

    if source_cell.font:
        destination_cell.font = copy(source_cell.font)

    if source_cell.fill:
        destination_cell.fill = copy(source_cell.fill)

    if source_cell.border:
        destination_cell.border = copy(source_cell.border)

    if source_cell.alignment:
        destination_cell.alignment = copy(source_cell.alignment)

    if source_cell.protection:
        destination_cell.protection = copy(source_cell.protection)

    if source_cell.comment:
        destination_cell.comment = copy(source_cell.comment)

    if source_cell.hyperlink:
        destination_cell._hyperlink = copy(source_cell.hyperlink)


def copy_row(
    source_worksheet,
    destination_worksheet,
    source_row_number: int,
    destination_row_number: int,
    maximum_column: int,
      ) -> None:
    """Copy one complete row from the source sheet to the output sheet."""
    for column_number in range(1, maximum_column + 1):
        source_cell = source_worksheet.cell(
            row=source_row_number,
            column=column_number,
        )

        destination_cell = destination_worksheet.cell(
            row=destination_row_number,
            column=column_number,
        )

        copy_cell(source_cell, destination_cell)

    source_height = source_worksheet.row_dimensions[source_row_number].height

    if source_height is not None:
        destination_worksheet.row_dimensions[
            destination_row_number
        ].height = source_height


def copy_column_widths(
    source_worksheet,
    destination_worksheet,
    maximum_column: int,
     ) -> None:
    """Copy source column widths and visibility settings."""
    for column_number in range(1, maximum_column + 1):
        column_letter = get_column_letter(column_number)

        source_dimension = source_worksheet.column_dimensions[column_letter]
        destination_dimension = destination_worksheet.column_dimensions[
            column_letter
        ]

        if source_dimension.width is not None:
            destination_dimension.width = source_dimension.width

        destination_dimension.hidden = source_dimension.hidden
        destination_dimension.bestFit = source_dimension.bestFit
        destination_dimension.outlineLevel = source_dimension.outlineLevel


def split_excel_file() -> list[Path]:
    """Perform the complete Excel splitting process."""
    input_path = Path(INPUT_FILE).expanduser().resolve()
    output_path = Path(OUTPUT_FOLDER).expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input Excel file not found: {input_path}")

    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("The input file must be an .xlsx or .xlsm file.")

    keep_vba = input_path.suffix.lower() == ".xlsm"

    source_workbook = load_workbook(
        filename=input_path,
        data_only=False,
        keep_vba=keep_vba,
    )

    if SHEET_NAME not in source_workbook.sheetnames:
        available_sheets = ", ".join(source_workbook.sheetnames)

        raise ValueError(
            f'Worksheet "{SHEET_NAME}" was not found. '
            f"Available worksheets: {available_sheets}"
        )

    source_worksheet = source_workbook[SHEET_NAME]

    header_row, selected_column = find_header(
        source_worksheet,
        COLUMN_NAME,
    )

    maximum_column = find_last_used_column(
        source_worksheet,
        header_row,
    )

    grouped_rows: OrderedDict[
        tuple[str, Any],
        dict[str, Any],
    ] = OrderedDict()

    for row_number in range(header_row + 1, source_worksheet.max_row + 1):
        distinct_value = source_worksheet.cell(
            row=row_number,
            column=selected_column,
        ).value

        is_blank = (
            distinct_value is None
            or (
                isinstance(distinct_value, str)
                and not distinct_value.strip()
            )
        )

        if is_blank and not INCLUDE_BLANK_VALUES:
            continue

        key = value_key(distinct_value)

        if key not in grouped_rows:
            grouped_rows[key] = {
                "value": distinct_value,
                "row_numbers": [],
            }

        grouped_rows[key]["row_numbers"].append(row_number)

    if not grouped_rows:
        raise ValueError(
            f'No data was found below header row {header_row} '
            f'for column "{COLUMN_NAME}".'
        )

    output_path.mkdir(parents=True, exist_ok=True)

    used_filenames: set[str] = set()
    created_files: list[Path] = []

    output_extension = ".xlsm" if keep_vba else ".xlsx"

    print(f'Detected header "{COLUMN_NAME}" at row {header_row}.')
    print(f"Found {len(grouped_rows)} distinct value(s).\n")

    for group in grouped_rows.values():
        distinct_value = group["value"]
        matching_row_numbers = group["row_numbers"]

        output_workbook = Workbook()
        output_worksheet = output_workbook.active

        # Excel worksheet names can contain a maximum of 31 characters.
        output_worksheet.title = source_worksheet.title[:31]

        copy_column_widths(
            source_worksheet,
            output_worksheet,
            maximum_column,
        )

        # The detected header row becomes row 1 in every output file.
        copy_row(
            source_worksheet=source_worksheet,
            destination_worksheet=output_worksheet,
            source_row_number=header_row,
            destination_row_number=1,
            maximum_column=maximum_column,
        )

        # Copy every row belonging to the current distinct value.
        for output_row_number, source_row_number in enumerate(
            matching_row_numbers,
            start=2,
        ):
            copy_row(
                source_worksheet=source_worksheet,
                destination_worksheet=output_worksheet,
                source_row_number=source_row_number,
                destination_row_number=output_row_number,
                maximum_column=maximum_column,
            )

        output_worksheet.freeze_panes = "A2"

        output_worksheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(maximum_column)}"
            f"{len(matching_row_numbers) + 1}"
        )

        safe_name = create_safe_filename(distinct_value)
        unique_name = create_unique_filename(
            safe_name,
            used_filenames,
        )

        destination_file = output_path / (
            f"{unique_name}{output_extension}"
        )

        output_workbook.save(destination_file)
        created_files.append(destination_file)

        print(
            f"Created: {destination_file.name} "
            f"({len(matching_row_numbers)} row(s))"
        )

    return created_files


def main() -> None:
    try:
        created_files = split_excel_file()

        print(
            f"\nCompleted successfully. "
            f"{len(created_files)} Excel file(s) created."
        )
        print(f"Output folder: {Path(OUTPUT_FOLDER).resolve()}")

    except Exception as error:
        print(f"\nError: {error}")
        raise


if __name__ == "__main__":
    main()
