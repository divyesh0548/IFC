"""
Convert an Excel sheet (sub-process + risk description columns) into risk category JSON.

Edit the configuration variables below, then run:
    python JSON_formatter.py

Dependencies:
    pip install openpyxl pyxlsb
"""

from __future__ import annotations

import json
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xlsb"}

# --- Configuration: edit these before running ---
EXCEL_PATH = SCRIPT_DIR / "Rubamin_Capex IFC_FY 2025-26_Testing sheet.xlsx"
SHEET_NAME = "IFC"
SUB_PROCESS_COLUMN = "Sub-Process"
RISK_DESCRIPTION_COLUMN = "Risks Event "
OUTPUT_DIR = SCRIPT_DIR

BUSINESS_PROCESS = ""
BUSINESS_PROCESS_CODE = ""
COMPANIES_INCLUDED: list[str] = []
# ------------------------------------------------


def _cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_matches_keywords(header: str, *keywords: str) -> bool:
    normalized = header.casefold()
    return all(keyword.casefold() in normalized for keyword in keywords)


def _detect_sub_process_column(headers: list[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header and _header_matches_keywords(header, "sub", "process"):
            return idx
    return None


def _detect_risk_description_column(headers: list[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header and _header_matches_keywords(header, "risk", "description"):
            return idx
    return None


def _resolve_column_indices(
    headers: list[str],
    sub_process_column: str,
    risk_description_column: str,
) -> tuple[int | None, int | None, str | None, str | None, bool]:
    """
    Resolve column indices from configured names, with keyword fallback per column.

    Returns:
        sub_idx, risk_idx, resolved_sub_name, resolved_risk_name, used_fallback
    """
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    used_fallback = False

    if sub_process_column in header_index:
        sub_idx = header_index[sub_process_column]
        resolved_sub_name = sub_process_column
    else:
        sub_idx = _detect_sub_process_column(headers)
        resolved_sub_name = headers[sub_idx] if sub_idx is not None else None
        if sub_idx is not None:
            used_fallback = True

    if risk_description_column in header_index:
        risk_idx = header_index[risk_description_column]
        resolved_risk_name = risk_description_column
    else:
        risk_idx = _detect_risk_description_column(headers)
        resolved_risk_name = headers[risk_idx] if risk_idx is not None else None
        if risk_idx is not None:
            used_fallback = True

    return sub_idx, risk_idx, resolved_sub_name, resolved_risk_name, used_fallback


def _pyxlsb_row_to_values(row) -> list:
    if not row:
        return []
    max_col = max(cell.c for cell in row)
    values = [None] * (max_col + 1)
    for cell in row:
        values[cell.c] = cell.v
    return values


def _iter_openpyxl_rows(excel_path: Path, sheet_name: str) -> Iterator[list]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
        )
    sheet = workbook[sheet_name]
    try:
        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _iter_pyxlsb_rows(excel_path: Path, sheet_name: str) -> Iterator[list]:
    with open_xlsb_workbook(excel_path) as workbook:
        if sheet_name not in workbook.sheets:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheets)}"
            )
        with workbook.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                yield _pyxlsb_row_to_values(row)


def _iter_sheet_rows(excel_path: Path, sheet_name: str) -> Iterator[list]:
    suffix = excel_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        yield from _iter_openpyxl_rows(excel_path, sheet_name)
    elif suffix == ".xlsb":
        yield from _iter_pyxlsb_rows(excel_path, sheet_name)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def _extract_risk_combinations(
    rows: Iterator[list],
    sub_process_column: str,
    risk_description_column: str,
) -> tuple[list[dict[str, str]], dict[str, str]]:
    sub_idx: int | None = None
    risk_idx: int | None = None
    resolved_columns: dict[str, str] = {}
    used_fallback = False
    last_headers: list[str] = []

    for row in rows:
        headers = [_cell_value(cell) for cell in row]
        if not any(headers):
            continue

        last_headers = headers
        sub_idx, risk_idx, resolved_sub, resolved_risk, row_fallback = _resolve_column_indices(
            headers,
            sub_process_column,
            risk_description_column,
        )
        if sub_idx is not None and risk_idx is not None:
            resolved_columns = {
                "sub_process_column": resolved_sub or "",
                "risk_description_column": resolved_risk or "",
            }
            used_fallback = row_fallback
            break

    if sub_idx is None or risk_idx is None:
        available = ", ".join(h for h in last_headers if h) or "(none)"
        raise ValueError(
            f"Could not find sub-process and risk-description columns. "
            f"Configured: '{sub_process_column}', '{risk_description_column}'. "
            f"Last scanned row columns: {available}"
        )

    if used_fallback:
        print(
            "Column fallback used - detected: "
            f"sub_process='{resolved_columns['sub_process_column']}', "
            f"risk_description='{resolved_columns['risk_description_column']}'"
        )

    risk_combinations: list[dict[str, str]] = []
    for row in rows:
        sub_process = _cell_value(row[sub_idx] if sub_idx < len(row) else None)
        risk_description = _cell_value(row[risk_idx] if risk_idx < len(row) else None)
        if not sub_process and not risk_description:
            continue
        risk_combinations.append(
            {
                "sub_process": sub_process,
                "risk_description": risk_description,
            }
        )

    return risk_combinations, resolved_columns


def excel_to_risk_json(
    excel_path: str | Path,
    sheet_name: str,
    sub_process_column: str,
    risk_description_column: str,
    *,
    output_dir: str | Path | None = None,
    business_process: str = "",
    business_process_code: str = "",
    companies_included: list[str] | None = None,
) -> Path:
    """
    Read sub-process and risk-description columns from an Excel sheet and write JSON.

    Supports .xlsx, .xlsm (openpyxl) and .xlsb (pyxlsb).

    Returns:
        Path to the written JSON file (same base name as the Excel file).
    """
    excel_path = Path(excel_path).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if excel_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("Only .xlsx, .xlsm, and .xlsb files are supported.")

    output_dir = Path(output_dir).resolve() if output_dir else SCRIPT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{excel_path.stem}.json"

    risk_combinations, _resolved_columns = _extract_risk_combinations(
        _iter_sheet_rows(excel_path, sheet_name),
        sub_process_column,
        risk_description_column,
    )

    payload = {
        "business_process": business_process,
        "business_process_code": business_process_code,
        "companies_included": companies_included or [],
        "risk_combinations": risk_combinations,
    }

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")

    return output_path


if __name__ == "__main__":
    output_path = excel_to_risk_json(
        excel_path=EXCEL_PATH,
        sheet_name=SHEET_NAME,
        sub_process_column=SUB_PROCESS_COLUMN,
        risk_description_column=RISK_DESCRIPTION_COLUMN,
        output_dir=OUTPUT_DIR,
        business_process=BUSINESS_PROCESS,
        business_process_code=BUSINESS_PROCESS_CODE,
        companies_included=COMPANIES_INCLUDED,
    )
    count = len(json.loads(output_path.read_text(encoding="utf-8"))["risk_combinations"])
    print(f"Wrote {count} risk combination(s) to {output_path}")
